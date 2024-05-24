import pyperclip
import re
import os
import subprocess
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.alignment import Alignment
from openpyxl.utils import *
from datetime import date
from datetime import datetime

########## FUNCTIONS ############
def setupIPSpreadSheet(ip_list):
    """I need to add commentss here""" 
  
    #Creating an instance of openpyxl
    my_workbook = openpyxl.Workbook()

    #Creating a workbook, naming it, and creating an additional sheet
    curr_sheet = my_workbook.active
    curr_sheet.title = "IP Addresses"
    
    curr_sheet["A1"] = "IP addresses"

  
    curr_sheet["A1"] = "IP's"

    curr_sheet["A1"].font = Font(b=True, color="000080")
    curr_sheet["A1"].alignment = Alignment(horizontal="center")

    row_count = 2
    for item in ip_list:
        col_count = 1
        curr_sheet.cell(row=row_count, column=col_count).value = item
        curr_sheet.cell(row=row_count, column=col_count).alignment = Alignment(horizontal="center")
        row_count += 1

    
    curr_sheet.freeze_panes = "A1"
    curr_sheet.column_dimensions["A"].width = 20
  
    #Save dat shisshhhhh
    my_workbook.save("C:\\IP_Info\\" + str(current_date) + "\\IP_Addresses_" + str(shorter_time) + ".xlsx")

    return shorter_time



def displayIPData():
    """I need to add commentss here"""
  
    allData = str(pyperclip.paste())
    ipList = findAllIPs(allData)

    output_string = ""
    count = 0
    while count < len(ipList):
        output_string +=  ipList[count] + "\n"
        count += 1
    
    return output_string


def ipListData():
    """I need to add commentss here"""
  
    allData = str(pyperclip.paste())
    ipList = findAllIPs(allData)
    
    return ipList


def findAllIPs(passed_string):
    """I need to add commentss here"""
  
    ipRegEx = re.compile(r'(?:\d{1,3}\.){3}\d{1,3}')

    ipList = ipRegEx.findall(passed_string)

    ip_return_list = []
    for ip in ipList:
        temp_ip = ""
        for item in ip:
            temp_ip += str(item)
        ip_return_list.append(ip)


    return ip_return_list




#Gets thee date and time.
current_date = date.today()
current_time = datetime.now().time()
shorter_time = current_time.strftime("%H-%S")

#Creates directories for excel files
os.chdir("C:\\")
try:
    os.makedirs("C:\\IP_Info\\")
    os.makedirs("C:\\IP_Info\\" + str(current_date))

except FileExistsError:
    pass

#This is a useless while loop
program_running = True
while program_running:

    #Collect Data and put into list
    new_data = displayIPData()
    ip_list = ipListData()

    #Copies to clipboard
    pyperclip.copy(new_data)

    #Creates excel spreadsheet and saves to c:\
    excel_time = setupIPSpreadSheet(ip_list)

    #Put file paths into variables
    excel_file_path = "C:\\IP_Info\\" + str(current_date) + "\\IP_Addresses_" + str(excel_time) + ".xlsx"
    notepad_file_path = "C:\\IP_Info\\" +  str(current_date) + "\\IP_Addresses_" + str(shorter_time) + ".txt"

    #Write Notepad file
    try:
        with open(notepad_file_path, 'w') as text_file:
            text_file.write(new_data)
    except:
        pass
    
    #Opens Excel
    subprocess.Popen(['start', '', excel_file_path], shell=True)

    program_running = False
